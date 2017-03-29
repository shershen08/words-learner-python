import openpyxl

import os
import sys
from time import gmtime, strftime

filename_xslx = 'words.xlsx'

def set_filename(name):
    global filename_xslx
    filename_xslx = name

def get_file_and_wb():
    try:
        wb = openpyxl.load_workbook(filename_xslx)
        return wb
    except OSError as e:
        print('Error: no file %s found.' % filename_xslx)
        sys.exit(0)

def read_language_pair ():
    sheet = get_file_and_wb().get_sheet_by_name('words')
    first_row = sheet[1]
    return first_row[0].value, first_row[1].value

def read_words_list (from_pos=0, to_pos=100):
    sheet = get_file_and_wb().get_sheet_by_name('words')
    return sheet

def get_total_lines(active_sheet):
    index = 0
    for i, row in enumerate(active_sheet.rows):
        index = i + 1
        if active_sheet.cell(column=2, row=index).value:
            pass

    return index + 1


def write_stats (words_numer, success_ratio):
    wb = get_file_and_wb()
    sheet = wb.get_sheet_by_name('results')
    test_itmestamp = strftime("%A, %d %b %H:%M", gmtime())
    next_index = get_total_lines(sheet)

    sheet.cell(column=1, row=next_index).value = test_itmestamp
    sheet.cell(column=2, row=next_index).value = words_numer
    sheet.cell(column=3, row=next_index).value = success_ratio
    wb.save(filename=filename_xslx)
